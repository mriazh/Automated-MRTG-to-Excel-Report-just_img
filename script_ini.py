import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage
import io

# ========== KONFIGURASI ==========
FOLDER_DATA = "MRTG-Data"
TEMPLATE_FILE = "MENTAHAN FORMAT DAILY MRTG.xlsx"
OUTPUT_FILE = "Daily_Report_Complete.xlsx"
DAFTAR_FILE = "list_mrtg_data_to_docx.txt"
MAPPING_FILE = "sid_image-position-excel.txt"  # file mapping area

# ========== BACA MAPPING AREA ==========
def baca_mapping(filepath):
    """
    Baca file mapping format:
    SID : 4700001-0021497479
    -> B12-L23
    SID : 4700001-0020265222
    -> N12-X23
    ...
    Return dictionary {id: (start_cell, end_cell)} 
    start_cell = (row, col), end_cell = (row, col)
    """
    mapping = {}
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith('SID : '):
            # extract id
            id_raw = line.replace('SID : ', '').strip()
            # Hapus kurung jika ada (misal (2897) 4703277-...)
            # Kita simpan id asli apa adanya, nanti untuk pencarian kita akan pakai substring
            # Tapi untuk mapping key, kita gunakan id yang sudah dibersihkan (tanpa kurung dan spasi ekstra)
            # Bersihkan: ambil angka setelah kurung jika ada
            id_clean = re.sub(r'^\(\d+\)\s*', '', id_raw)  # hapus (2897) dll
            i += 1
            if i < len(lines) and lines[i].startswith('->'):
                range_str = lines[i][2:].strip()  # ambil "B12-L23"
                # Parse range
                start, end = range_str.split('-')
                start_col = column_index_from_string(re.match(r'[A-Z]+', start).group())
                start_row = int(re.search(r'\d+', start).group())
                end_col = column_index_from_string(re.match(r'[A-Z]+', end).group())
                end_row = int(re.search(r'\d+', end).group())
                mapping[id_clean] = ((start_row, start_col), (end_row, end_col))
                i += 1
            else:
                i += 1
        else:
            i += 1
    return mapping

# ========== BACA DAFTAR ITEM ==========
def baca_daftar(filepath):
    items = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split('.', 1)
            if len(parts) != 2:
                continue
            nomor = parts[0].strip()
            rest = parts[1].strip()
            if rest.startswith('SID : '):
                tipe = 'SID'
                id_val = rest.replace('SID : ', '').strip()
            elif rest.startswith('Graph-title : '):
                tipe = 'Graph-title'
                id_val = rest.replace('Graph-title : ', '').strip()
            else:
                continue
            items.append((nomor, tipe, id_val))
    return items

# ========== HITUNG UKURAN AREA DALAM PIKEL ==========
def get_area_size_pixels(sheet, start_row, start_col, end_row, end_col):
    """
    Menghitung lebar (pixel) dan tinggi (pixel) dari area yang dibatasi start_row,start_col sampai end_row,end_col.
    Lebar kolom: 1 unit = 7.5 pixel (kira-kira, sesuai openpyxl default)
    Tinggi baris: 1 point = 1.333 pixel (karena 1 point = 1/72 inch, 1 pixel = 1/96 inch -> 1 point = 1.333 pixel)
    """
    total_width = 0
    for col in range(start_col, end_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        col_width = sheet.column_dimensions[col_letter].width
        if col_width is None:
            col_width = 8.43  # default Excel
        total_width += col_width * 7.4  # konversi ke pixel (estimasi)
    
    total_height = 0
    for row in range(start_row, end_row + 1):
        row_height = sheet.row_dimensions[row].height
        if row_height is None:
            row_height = 15  # default Excel dalam point
        total_height += row_height * 1.333  # konversi ke pixel
    
    return total_width, total_height

# ========== RESIZE GAMBAR KE UKURAN TERTENTU ==========
def resize_image_to_size(image_path, target_width, target_height):
    """Resize gambar proporsional? Atau stretch? Kita pilih stretch agar pas di area."""
    with PILImage.open(image_path) as img:
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        # Stretch ke target width/height
        img_resized = img.resize((int(target_width), int(target_height)), PILImage.Resampling.LANCZOS)
        output = io.BytesIO()
        img_resized.save(output, format='PNG')
        output.seek(0)
        return output

# ========== TAMBAHKAN GAMBAR KE SHEET ==========
def tambah_gambar_di_area(sheet, image_path, start_row, start_col, end_row, end_col):
    """Resize gambar sesuai area dan letakkan dengan anchor di start_cell."""
    try:
        # Hitung ukuran area dalam pixel
        width_px, height_px = get_area_size_pixels(sheet, start_row, start_col, end_row, end_col)
        if width_px <= 0 or height_px <= 0:
            print(f"    Ukuran area tidak valid: {width_px}x{height_px}")
            return False
        # Resize gambar
        img_bytes = resize_image_to_size(image_path, width_px, height_px)
        img = Image(img_bytes)
        # Anchor ke sel kiri atas
        cell_anchor = openpyxl.utils.get_column_letter(start_col) + str(start_row)
        img.anchor = cell_anchor
        # Set ukuran gambar (opsional, karena sudah di-resize sesuai target)
        # openpyxl akan menggunakan ukuran dari gambar itu sendiri
        sheet.add_image(img)
        return True
    except Exception as e:
        print(f"    Gagal tambah gambar: {e}")
        return False

# ========== PROSES SATU TANGGAL ==========
def proses_tanggal(wb, tanggal_str, items, mapping):
    hari = int(tanggal_str[6:8])
    sheet_name = f"{hari:02d}"
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet {sheet_name} tidak ditemukan")
        return
    sheet = wb[sheet_name]
    print(f"  Memproses sheet {sheet_name}...")
    
    for nomor, tipe, id_val in items:
        # Bersihkan id_val untuk dicocokkan dengan mapping (hapus kurung jika ada)
        id_clean = re.sub(r'^\(\d+\)\s*', '', id_val)
        if id_clean not in mapping:
            print(f"    Peringatan: ID '{id_clean}' tidak ditemukan di mapping")
            continue
        (start_row, start_col), (end_row, end_col) = mapping[id_clean]
        
        # Tentukan path gambar
        if tipe == 'SID':
            nama_file = f"MRTG_{id_val}.png"
        else:
            nama_file = f"MRTG_{id_val}_{tanggal_str}.png"
        path_gambar = os.path.join(FOLDER_DATA, tanggal_str, nama_file)
        if not os.path.exists(path_gambar):
            # Fallback: cari file dengan awalan MRTG_{id_val}
            folder_tgl = os.path.join(FOLDER_DATA, tanggal_str)
            if os.path.exists(folder_tgl):
                for f in os.listdir(folder_tgl):
                    if f.startswith(f"MRTG_{id_val}") and f.endswith(".png"):
                        path_gambar = os.path.join(folder_tgl, f)
                        break
        if not os.path.exists(path_gambar):
            print(f"    Gambar tidak ditemukan: {path_gambar}")
            continue
        
        print(f"    Menambahkan gambar untuk {tipe} {id_val} di area {start_row},{start_col} - {end_row},{end_col}")
        tambah_gambar_di_area(sheet, path_gambar, start_row, start_col, end_row, end_col)

# ========== MAIN ==========
def main():
    print("=" * 60)
    print("AUTOMATED MRTG TO EXCEL - DENGAN MAPPING AREA")
    print("=" * 60)
    
    # Baca mapping area
    if not os.path.exists(MAPPING_FILE):
        print(f"File mapping '{MAPPING_FILE}' tidak ditemukan!")
        return
    mapping = baca_mapping(MAPPING_FILE)
    print(f"Mapping berisi {len(mapping)} entri.")
    
    # Baca daftar item
    items = baca_daftar(DAFTAR_FILE)
    print(f"Daftar berisi {len(items)} item.")
    
    # Dapatkan folder tanggal
    if not os.path.exists(FOLDER_DATA):
        print(f"Folder '{FOLDER_DATA}' tidak ditemukan!")
        return
    tanggal_list = [d for d in os.listdir(FOLDER_DATA) if os.path.isdir(os.path.join(FOLDER_DATA, d)) and d.isdigit() and len(d) == 8]
    tanggal_list.sort()
    print(f"Ditemukan {len(tanggal_list)} folder tanggal: {tanggal_list[:5]}...")
    
    # Load template
    if not os.path.exists(TEMPLATE_FILE):
        print(f"Template '{TEMPLATE_FILE}' tidak ditemukan!")
        return
    wb = load_workbook(TEMPLATE_FILE)
    print("Template berhasil dimuat.")
    
    # Proses setiap tanggal
    for tgl in tanggal_list:
        print(f"\nMemproses tanggal: {tgl}")
        proses_tanggal(wb, tgl, items, mapping)
    
    # Simpan
    wb.save(OUTPUT_FILE)
    print("\n" + "=" * 60)
    print(f"🎉 SELESAI! File Excel disimpan sebagai: {OUTPUT_FILE}")
    print("=" * 60)

if __name__ == "__main__":
    import openpyxl
    main()